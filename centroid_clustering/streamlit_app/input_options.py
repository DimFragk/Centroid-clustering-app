import pandas as pd
import numpy as np
import streamlit as st
from openpyxl import load_workbook
from sklearn.datasets import make_classification, make_blobs

import math
from io import BytesIO
from os.path import isfile

from utils.general_functions import list_of_names_in_list
from utils.math_functions import linspace_list
from utils.pandas_functions import check_col_dtype

from .app_classes import PAMInput
from .result_pages import show_3d_plots

from .. import clustering_selection as pam


def pam_input_from_single_df(samples_df: pd.DataFrame):
    labels = None
    centers = None
    cols = list(samples_df.columns)
    label_col = list_of_names_in_list(["Target", "target", "Labels", "labels", "lb"], cols)

    c1, c2, c3 = st.columns((4, 1, 1), vertical_alignment="bottom")

    if label_col:
        str_col_list = check_col_dtype(samples_df, str)
        rem_cols = c1.multiselect(
            label="Selected features",
            options=cols,
            default=list(filter(lambda x: x != label_col and x not in str_col_list, cols))
        )
        samples = samples_df[rem_cols]

        labels = samples_df[label_col]
        c2.markdown(":blue-background[:blue[Target labels detected]]")
    else:
        samples = df_m_slt_cols(samples_df, st_cont=c1)
        lb_col_name = c2.text_input(label="Label columns name")
        if lb_col_name in cols:
            labels = samples_df[lb_col_name]
        elif len(lb_col_name) == 1:
            try:
                if int(lb_col_name) in cols:
                    labels = samples_df[int(lb_col_name)]
            except ValueError:
                pass

    diss = True
    if labels is not None:
        if labels.values[0] in samples_df.index and labels.values[-1] in samples_df.index:
            diss = False

    if c3.toggle(label="Medoids from samples", value=not diss, disabled=diss):
        centers = samples.loc[labels.unique()]

    return samples, labels, centers


def df_m_slt_cols(samples_df, st_cont=None):
    if st_cont is None:
        st_cont = st.container()

    cols = list(samples_df.columns)
    str_col_list = check_col_dtype(samples_df, str)
    rem_cols = st_cont.multiselect(
        label="Selected features",
        options=cols,
        default=list(filter(lambda x: x not in str_col_list, cols))
    )
    return samples_df[rem_cols]


@st.fragment
def from_file_to_pam_input_obj(obj: pd.DataFrame | BytesIO, f_name):
    if isinstance(obj, pd.DataFrame):
        samples, labels, centers = pam_input_from_single_df(obj)
        submit_data_gen(
            name=f_name,
            points=samples,
            labels=labels,
            centers=centers,
            st_name=False
        )
        return

    my_workbook = load_workbook(obj)
    sheet_names = my_workbook.sheetnames

    labels = None
    centers = None

    if len(sheet_names) == 1:
        samples_df = pd.read_excel(
            obj,
            header=0,
            index_col=0
        )
        samples, labels, centers = pam_input_from_single_df(samples_df)
        submit_data_gen(
            name=f_name,
            points=samples,
            labels=labels,
            centers=centers,
            st_name=False
        )
        return

    opt_sheet_name = list_of_names_in_list(["Options", "options", "Opt", "opt"], sheet_names)
    n_samples = None
    n_features = None
    n_clusters = None
    sample_cps = None
    if opt_sheet_name:
        opt_df = pd.read_excel(
            obj,
            sheet_name=opt_sheet_name,
            header=0,
            index_col=0,
            usecols=[0, 1]
        ).squeeze()
        n_samples = opt_df["Samples"]
        n_features = opt_df["Features"]
        n_clusters = opt_df["Clusters"]
        sample_cps = opt_df["Sample centers"]

    s_name = list_of_names_in_list(
        ["Samples", "samples", "S", "s", "Data", "data", "Points", "points"], sheet_names
    )
    s_name = st.selectbox(
        label="Samples data sheet name", options=sheet_names, index=sheet_names.index(s_name) if s_name else None
    )
    samples_df = pd.read_excel(
        obj,
        sheet_name=s_name,
        header=0,
        index_col=0
    )

    c_names = list_of_names_in_list(
        ["Centers", "centers", "CPS", "Cps", "cps", "C", "c"], sheet_names
    )
    c_names = st.selectbox(
        label="Centers data sheet name",
        options=sheet_names + [False], index=sheet_names.index(c_names) if c_names else len(sheet_names)
    )
    centers_df = None
    if c_names:
        centers_df = pd.read_excel(
            obj,
            sheet_name=c_names,
            header=0,
            index_col=0
        )

    lb_name = list_of_names_in_list(
        ["Target", "target", "Labels", "labels", "lb"], sheet_names
    )

    if lb_name is None:
        lb_name = st.selectbox(label="Labels exist?", options=["In sample data", "No"] + sheet_names)
        if lb_name == "No":
            samples = df_m_slt_cols(samples_df)
            submit_data_gen(
                name=f_name,
                points=samples,
                labels=pd.Series(1, index=samples_df.index),
                st_name=False
            )
            return
        elif lb_name == "In sample data":
            samples, labels, centers = pam_input_from_single_df(samples_df)

            submit_data_gen(
                name=f_name,
                points=samples,
                labels=labels if labels is not None else pd.Series(1, index=samples_df.index),
                centers=centers_df if centers_df is not None else centers,
                st_name=False
            )
            return

    samples = df_m_slt_cols(samples_df)

    labels = pd.read_excel(
        obj,
        sheet_name=lb_name,
        header=0,
        index_col=0,
        # usecols=1
    ).squeeze()

    submit_data_gen(
        name=f_name,
        points=samples,
        labels=labels,
        centers=centers_df,
        st_name=False
    )


@st.fragment
def create_random_data_with_sklearn_make_cl():
    c1, c2, c3 = st.columns(3)
    """
    name = c0.text_input(
        label="Name of data gen", value=f"sklearn make classification: v{len(st.session_state.method_input)+1}"
    )
    """
    name = "sklearn make classification"
    n_s = c1.number_input(
        label="Number of samples", value=1000, min_value=5, step=1, key=f"number_input_n_s_sklearn_make_cl"
    )
    number_of_cl = c2.number_input(
        label="Number of classes (unique labels)",
        value=5, min_value=2, max_value=n_s, key=f"slider_n_c_sklearn_make_cl"
    )
    number_of_features = c3.number_input(label="Number of features", value=5, min_value=2, max_value=50, step=1)

    c1, c2, c31, c32, c4, c5 = st.columns((4, 4, 1, 1, 4, 4), vertical_alignment="center", gap="large")

    max_n_cl = int((2 ** number_of_features)/number_of_cl)
    n_clusters_per_class = c4.slider(
        label="Number of clusters per class",
        min_value=1, max_value=max_n_cl, value=1,
    )
    class_sep = c5.slider(
        label="Separation of clusters/Classes",
        min_value=0.1, max_value=5.0, step=0.01, value=1.0,
        help="""
        The factor multiplying the hypercube size. 
        Larger values spread out the clusters/classes and make the classification task easier.
        """
    )
    hypercube = c31.toggle(
        label="H", value=True, help="""
        If True, the clusters are put on the vertices of a hypercube. 
        If False, the clusters are put on the vertices of a random polytope.
        """
    )
    r_state = c32.select_slider(
        label="R", options=[None]+list(range(1, 21)), value=1, help="""
            Determines random number generation for dataset creation. 
            Pass an int for reproducible output across multiple function calls.
            """
    )

    # number_of_cl * n_clusters_per_class <= 2 ** n_formative
    min_val = math.ceil(math.log2((number_of_cl * n_clusters_per_class)))
    diss = min_val >= number_of_features
    n_formative = c1.slider(
        label="Informative features",
        min_value=min_val,
        max_value=number_of_features,
        value=number_of_features if not diss else number_of_features + 1,
        disabled=diss,
        help="""
            The number of informative features. 
            Each class is composed of a number of gaussian clusters
            each located around the vertices of a hypercube
            in a subspace of dimension n_informative. 
            For each cluster, informative features are drawn independently from N(0, 1)
            and then randomly linearly combined within each cluster in order to add covariance. 
            The clusters are then placed on the vertices of the hypercube.
            """
    )
    dis = number_of_features <= n_formative
    if dis:
        n_formative = number_of_features

    n_redundant = c2.slider(
        label="Redundant features",
        min_value=0, max_value=number_of_features - n_formative if not dis else 1, value=0, disabled=dis,
        help="""
            The number of redundant features. 
            These features are generated as random linear combinations of the informative features.
            """
    )
    if n_formative + n_redundant > number_of_features:
        n_redundant = 0

    points, labels = make_classification(
        n_samples=n_s,
        n_features=number_of_features,
        n_informative=n_formative,
        n_redundant=n_redundant,
        n_classes=number_of_cl,
        n_clusters_per_class=n_clusters_per_class,
        class_sep=class_sep,
        hypercube=hypercube,
        random_state=r_state
    )

    submit_data_gen(
        name=name,
        points=pd.DataFrame(points, columns=[f"Feature_{i}" for i in range(1, number_of_features + 1)]),
        labels=pd.Series(labels),
        st_name=True
    )


@st.fragment
def random_data_gen():
    points, labels, cps = pam.create_data_fit_for_clustering()
    """
    name = st.text_input(
        label="Name of data gen", value=f"default data gen: v{len(st.session_state.method_input) + 1}"
    )
    """
    name = "Default random data gen"
    print(points)
    print(labels)
    submit_data_gen(
        name=name,
        points=points,
        labels=labels,
        centers=cps,
        st_name=True
    )


@st.fragment
def custom_make_blobs():
    c1, c2 = st.columns(2)
    """
    name = c0.text_input(
        label="Name of data gen", value=f"Custom make blobs: v{len(st.session_state.method_input) + 1}"
    )
    """
    name = "Custom make blobs"
    number_of_blobs = c1.number_input(label="Number of 'make_blob' calls", value=3, min_value=1, step=1, max_value=5)
    number_of_features = c2.number_input(label="Number of features", value=5, min_value=2, max_value=50, step=1)
    """
    @dataclass
    class CallBlob:
        n_samples: int
        centers: int | list
        cluster_std: int | list
        center_box: tuple[float]
        random_state: int | None
    """
    feature_names = [f"Feature_{i}" for i in range(1, number_of_features+1)]

    data_list = []
    labels_list = []
    centers_list = []
    for i, col in enumerate(st.columns(number_of_blobs, gap="large")):
        with col:
            n_s = st.number_input(
                label="Number of samples", value=1000, min_value=5, step=1, key=f"number_input_n_s_{i}"
            )
            number_of_cl = st.number_input(
                label="Number of centers", value=5, min_value=1, max_value=n_s, key=f"slider_n_c_{i}"
            )
            n_c = number_of_cl
            c3, c4 = st.columns(2, vertical_alignment="center")
            if c3.toggle("Cluster samples in range", key=f"cl_samples_range{i}"):
                scale_factor = c4.slider(
                    label="Cluster size range scale",
                    min_value=0.1, max_value=1.0, step=0.05,
                    key=f"slider_mean_s_range{i}"
                )
                cl_samples = n_s / number_of_cl
                step = 2*cl_samples*scale_factor / (number_of_cl - 1)
                cl_samples_min = cl_samples*(1-scale_factor)
                n_s_list = [int(cl_samples_min + step*j) for j in range(number_of_cl)]
                print("pamp_line_1123")
                print(n_c)
                print(f"n_s ~= {sum(n_s_list)}")
                n_s_list[-1] += n_s - sum(n_s_list)
                n_s = n_s_list
                n_c = None
            else:
                st.write("")
                for j in range(5):
                    c4.text("")

            c3, c4 = st.columns(2, vertical_alignment="center")
            if c3.toggle("Different cluster std", key=f"toggle_cl_std{i}"):
                std_min_max = c4.slider(
                    label="Std min-max range",
                    min_value=0.01, max_value=3.0, step=0.01, value=(0.5, 1.5),
                    key=f"slider_std_range_1_{i}"
                )
                std_v = linspace_list(std_min_max[0], std_min_max[1], num=number_of_cl)
            else:
                std_v = c4.slider(
                    label="Std value",
                    min_value=0.01, max_value=3.0, step=0.01, value=0.8,
                    key=f"slider_std_range_2_{i}"
                )

            c_b = st.slider(label="Random centers box range", value=(-3.0, 3.0), key=f"slider_rand_center_box{i}")
            r_s = st.select_slider(label="Random state", options=[None] + list(range(20)), key=f"slider_rand_state{i}")

            data, labels, centers = make_blobs(
                n_samples=n_s,
                n_features=number_of_features,
                centers=n_c,
                cluster_std=std_v,
                center_box=c_b,
                shuffle=True,
                random_state=r_s,
                return_centers=True
            )
            data_list += [data]
            labels_list += [np.array([f"b({i+1}) : c{j+1}" for j in labels])]
            centers_list += [pd.DataFrame(
                centers, index=[f"b({i+1}) : c{j+1}" for j in range(len(centers))], columns=feature_names
            )]

    submit_data_gen(
        name=name,
        points=pd.DataFrame(np.vstack(data_list), columns=feature_names),
        labels=pd.Series(np.hstack(labels_list)),
        centers=pd.concat(centers_list),
        st_name=True
    )


def write_dataframe_to_excel(
        full_file_path,
        dataframe: pd.DataFrame = None,
        sheet_name: str = None,
        sheet_df_dict: dict[str, pd.DataFrame] = None,
        overwrite=False,
        start_row=0,
        start_col=0
):
    if sheet_df_dict is None:
        if not isinstance(dataframe, pd.DataFrame):
            print("No dataframe passed to object")
            return

        sheet_df_dict = {sheet_name: dataframe}
    else:
        if isinstance(dataframe, pd.DataFrame):
            if sheet_name is None:
                sheet_name = "new_sheet"

            if sheet_name in sheet_df_dict.keys():
                sheet_df_dict[f"{sheet_name}(1)"] = dataframe
            else:
                sheet_df_dict[sheet_name] = dataframe

    def excel_write(xl_wr):
        for sh_name, dframe in sheet_df_dict.items():
            dframe.to_excel(xl_wr, sheet_name=sh_name, startrow=start_row, startcol=start_col)

    def pd_xl_writer(mode="w", if_sheet_exists=None):
        return pd.ExcelWriter(full_file_path, mode=mode, if_sheet_exists=if_sheet_exists)

    if not isinstance(full_file_path, BytesIO) and isfile(full_file_path):
        if overwrite is False:
            with pd_xl_writer(mode="a", if_sheet_exists="new") as excel_wr:
                excel_write(excel_wr)
        elif overwrite is True:
            with pd_xl_writer(mode="a", if_sheet_exists="overlay") as excel_wr:
                excel_write(excel_wr)
    else:
        with pd_xl_writer(mode="w") as excel_wr:
            excel_write(excel_wr)
        # self.data_frame2wr_list += [dataframe]


def write_data_gen_to_buffer(points, labels, centers):
    buffer = BytesIO()
    write_dataframe_to_excel(
        full_file_path=buffer,
        sheet_df_dict={
            "Options": pd.Series(
                [
                    points.shape[0],
                    points.shape[1],
                    len(centers) if centers is not None else len(labels.unique()),
                    0 if centers is None else 1
                ],
                index=["Samples", "Features", "Clusters", "Sample centers"]
            ).to_frame(),
            "Samples": points,
            "Labels": labels,
            "Centers": centers if centers is not None else pd.DataFrame()
        }
    )
    return buffer


def submit_data_gen(name: str, points, labels, centers=None, st_name=False):
    show_data_gen(points, labels)

    if st_name:
        c01, c02 = st.columns((4, 10), vertical_alignment="top")
        # c1, c2, c3, c4 = st.columns((2, 2, 5, 8), vertical_alignment="bottom")
        input_name = c01.text_input(
            label="Name of data gen",
            value=f"{name}: v{len(st.session_state.method_input)+1}",
            label_visibility="collapsed",
            help="Give a name to the generated data"
        )
    else:
        input_name = name
        c01 = c02 = st.container()

    def submit_fn():
        print("pam_app_line_1420")
        print(name)
        res = PAMInput(
            f_name=input_name,
            data=points,
            target_labels=labels,
            center_points=centers
        )
        st.session_state.method_input["Setting up..."] = res

    c02.download_button(
        label="Download excel",
        data=write_data_gen_to_buffer(points, labels, centers),
        file_name=f"{input_name}.xlsx",
        mime='application/vnd.ms-excel',
        type="secondary",
        use_container_width=False   # st_name
    )

    if c01.button(label="Submit data", type="primary", on_click=submit_fn, use_container_width=False):
        st.rerun()
        print("pam_app_line_1432")
        print(input_name)


def show_data_gen(points, labels):
    c1, c2 = st.columns(2)
    with c1:
        show_3d_plots(data_df=points, labels_sr=labels, select_redux="PCA")
    with c2:
        show_3d_plots(data_df=points, labels_sr=labels, select_redux="LDA")


def create_random_data():
    options_fn_dict = {
        "Default 'make_blobs'": random_data_gen,
        "sklearn 'make_classification'": create_random_data_with_sklearn_make_cl,
        "custom make blobs": custom_make_blobs
    }
    type_slt = st.selectbox(label="Select type of random data", options=options_fn_dict.keys())

    res = options_fn_dict[type_slt]()

    if res is not None:
        if st.button(label="Submit data", type="primary"):
            return res

