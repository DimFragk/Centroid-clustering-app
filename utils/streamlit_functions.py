import pandas as pd
import numpy as np
import streamlit as st
from openpyxl import load_workbook
from io import BytesIO
import altair as alt
import plotly.express as px
import plotly.graph_objects as go

from . import general_functions as mf


def init_dec_key(father_key, direct_st_val=False):
    def partial_dec_key(func, **kwargs):
        st_val, st_key = dec_key(func, father_key=father_key, **kwargs)
        if direct_st_val:
            return st.session_state[st_key]

        return st_val, st_key

    return partial_dec_key


def dec_key(func, father_key=None, **kwargs):
    k_key = ""
    label = ""
    if 'key' in kwargs:
        k_key = kwargs.pop('key')
    if father_key is None:
        father_key = ""
    if 'label' in kwargs:
        label = kwargs['label']

    st_key = f"{father_key} {func.__name__} {k_key} {label}"
    return func(**kwargs, key=st_key), st_key


# @st.cache(allow_output_mutation=True)
def create_altair_m_line_plot_with(df_in_plt_form):
    df_cols_name = "variable" if df_in_plt_form.columns.name is None else df_in_plt_form.columns.name
    df_index_name = "index" if df_in_plt_form.index.name is None else df_in_plt_form.index.name

    data = df_in_plt_form.reset_index().melt(f"{df_index_name}")
    # print(data)
    # print(data.columns)
    x_lb = list(range(len(df_in_plt_form.index)))
    print(x_lb)
    alt_chart = alt.Chart(data).mark_line().encode(
        x=alt.X(f'{df_index_name}:T', axis=alt.Axis(title="Criteria values", values=x_lb, grid=True)),
        y=alt.Y('value:Q', scale=alt.Scale(domain=[0, 1]), axis=alt.Axis(title="Utility", grid=True)),
        color=f'{df_cols_name}:N'
    ).configure_axis(grid=True)

    return alt_chart


def plotly_h_bar_chart(histogram=True, axis=0):
    return lambda data_pd: st.plotly_chart(mf.set_up_h_bar_chart(data_pd, histogram, axis))


def st_latex_tab(text, size=5):
    return st.tabs([latext_size(text, size)])


def form_tabs(tab_names: list, size=4):
    return st.tabs([latext_size(name, size=size) for name in tab_names])


def latext_size(text, size=5):
    text_size_dict = [
        r"\Huge", r"\huge", r"\LARGE", r"\Large", r"\large",
        r"\normalsize", r"\small", r"\footnotesize", r"\scriptsize", r"\tiny"
    ]
    return fr"$\textsf{{{text_size_dict[size]} {text}}}$"


def st_selectbox(label, options_dict, **kwargs):
    return st.selectbox(
        label=label,
        options=list(options_dict.keys()),
        format_func=lambda x: options_dict.get(x),
        **kwargs
    )


# @st.cache(allow_output_mutation=True)
def create_altair_bar_charts_with_df_cols(dataframe, width=500, height=500):
    df_cols_name = "variable" if dataframe.columns.name is None else dataframe.columns.name
    df_index_name = "index" if dataframe.index.name is None else dataframe.index.name

    # print(dataframe)
    # print(dataframe.reset_index().melt(f"{df_index_name}"))
    data = pd.melt(dataframe.reset_index(), id_vars=[f"{df_index_name}"])
    # Horizontal stacked bar chart
    chart = (
        alt.Chart(data).mark_bar().encode(
            x=alt.X("value", type="quantitative", title=""),
            y=alt.Y(f"{df_index_name}", type="nominal", title=""),
            color=alt.Color(f"{df_cols_name}", type="nominal", title=""),
            order=alt.Order(f"{df_cols_name}", sort="descending"),
        ).properties(
            width=width,
            height=height
        )
    )
    return chart


# @st.cache
def convert_df_to_csv(df):
    # IMPORTANT: Cache the conversion to prevent computation on every rerun
    return df.to_csv().encode('utf-8')


# @st.cache
def return_dataframe_from_csv_or_excel(uploaded_file):
    if uploaded_file.name.split(".")[1] == "csv":
        return pd.read_csv(uploaded_file)
    else:
        return pd.read_excel(uploaded_file)


@st.cache_data
def load_workbook_from_bytes(bytes_io_file):
    return None if bytes_io_file is None else load_workbook(filename=bytes_io_file)     # openpyxl.load_workbook


def return_excel_workbook_with_st_file_uploader(label, label_visibility="collapsed"):
    bytes_io_file, file_name = rtn_file_in_bytes_io_with_st_file_uploader(label, label_visibility=label_visibility)
    my_workbook = load_workbook_from_bytes(bytes_io_file)
    return my_workbook, file_name


@st.cache_data
def return_file_in_bytes_io_from(st_uploaded_file):
    if st_uploaded_file is None:
        return None, None
    """
    bytes_io_file = None
    file_name = None
    if st_uploaded_file is not None:
        bytes_io_file = BytesIO(st_uploaded_file.getvalue())
        file_name = st_uploaded_file.name
    """
    bytes_io_file = BytesIO(st_uploaded_file.getvalue())
    file_name = st_uploaded_file.name
    return bytes_io_file, file_name


def rtn_file_in_bytes_io_with_st_file_uploader(label, label_visibility="collapsed"):
    uploaded_file = st.file_uploader(label=label, label_visibility=label_visibility)
    bytes_io_file, file_name = return_file_in_bytes_io_from(uploaded_file)
    return bytes_io_file, file_name


def add_mean_and_std(fig, data_sr: pd.Series, std=True, min_val=None, max_val=None, rect_text="", axis=0):
    data_mean = data_sr.mean()
    st_red = "#FF4B4B"
    st_green = "#35FFBA"
    st_blue = "#000D48"

    if axis == 0:
        fig.add_hline(
            y=data_mean, line_width=1, line_dash="dash", line_color=st_red
        )
    elif axis == 1:
        fig.add_vline(
            x=data_mean, line_width=1, line_dash="dash", line_color=st_red
        )

    if std is False and (min_val is None or max_val is None):
        return fig

    data_std = data_sr.std()
    if std is True:
        if axis == 0:
            fig.add_hrect(
                y0=data_mean-data_std, y1=data_mean+data_std, label=dict(
                    text="1 std",
                    textposition="top left",
                    font=dict(size=10)  # , family="Times New Roman"),
                ),
                fillcolor=st_blue,
                opacity=0.3,
                line_width=1,
            )
        elif axis == 1:
            fig.add_vrect(
                x0=data_mean - data_std, x1=data_mean + data_std, label=dict(
                    text="1 std",
                    textposition="top left",
                    font=dict(size=10)  # , family="Times New Roman"),
                ),
                fillcolor=st_blue,
                opacity=0.3,
                line_width=1,
            )

    if min_val is None or max_val is None:
        return fig

    if axis == 0:
        return fig.add_hrect(
            y0=min_val, y1=max_val, label=dict(
                text=f"{rect_text}",
                textposition="top right",
                font=dict(size=10)  # , family="Times New Roman"),
            ),
            fillcolor=st_green,
            opacity=0.1,
            line_width=1,
        )
    elif axis == 1:
        return fig.add_vrect(
            x0=min_val, x1=max_val, label=dict(
                text=f"{rect_text}",
                textposition="top right",
                font=dict(size=10)  # , family="Times New Roman"),
            ),
            fillcolor=st_green,
            opacity=0.1,
            line_width=1,
        )


def create_form_to_sidebar_if(s_state_bool, form_key, in_expander=False):
    """
    print(s_state_bool)
    if s_state_bool is True:
        with st.sidebar:
            if in_expander:
                with st.expander(label=form_key, expanded=True):
                    method_form = st.form(key=form_key)
            else:
                method_form = st.form(key=form_key)
    else:
        if in_expander:
            with st.expander(label=form_key, expanded=True):
                method_form = st.form(key=form_key)
        else:
            method_form = st.form(key=form_key)

    return method_form
    """
    with st.sidebar if s_state_bool is True else st.container():
        if in_expander:
            with st.expander(label=form_key, expanded=True):
                method_form = st.form(key=form_key)
        else:
            method_form = st.form(key=form_key)

    return method_form


def multiselect_submit(label, options, default, label_above=True, key=None, **kwargs):
    if f"mult_slt({label})" not in st.session_state.keys():
        st.session_state[f"mult_slt({label})"] = default

    with st.form(key=f"{key}-{label}"):
        slt = st.multiselect(
            label=label,
            options=options,
            label_visibility="visible" if label_above else "collapsed",
            default=default,
            key=key,
            **kwargs
        )
        if st.form_submit_button(label="apply" if label_above else label):
            st.session_state[f"mult_slt({label})"] = slt

    return st.session_state[f"mult_slt({label})"]


# @st.cache
def collection_counter_on_dataframe_columns(plot_df):
    return mf.collection_counter_on_dataframe_columns(plot_df)


# @st.cache
def prep_crit_df_for_plot(ij_crit_df, add_zero_point=True, explode_row=False):
    crit_df = ij_crit_df.applymap(lambda x: [0] + x) if add_zero_point is True else ij_crit_df

    def explode_sr(sr):
        return pd.DataFrame(sr.values.tolist(), index=sr.index).T

    if explode_row:
        return {crit_row: explode_sr(crit_df.loc[crit_row]) for crit_row in crit_df.index}

    else:
        """
            df_to_plot_dict = {}
            for col in crit_df:
                df_to_plot_dict[col] = pd.DataFrame(data=crit_df[col].values.tolist(), index=crit_df[col].index).T
            return df_to_plot_dict
        """
        return {crit_col: explode_sr(crit_df[crit_col]) for crit_col in crit_df.columns}


# @st.cache
def extract_p_opt_var_from(uta_obj, var_name, df=False):
    return uta_obj.extract_var_from_utils_obj_p_opt_dict(var_name, df=df)


def create_crit_multi_line_chart_from_df_columns(dict_of_plot_dfs: dict[str, pd.DataFrame]):
    for col, crit in zip((st.tabs(dict_of_plot_dfs.keys())), dict_of_plot_dfs):
        with col:
            df_col_group_dict, df_unique_values = collection_counter_on_dataframe_columns(dict_of_plot_dfs[crit])
            fig = go.Figure()
            for idx in df_unique_values.columns:
                col_rs = df_unique_values[idx]
                fig.add_trace(go.Scatter(
                    x=col_rs.index, y=col_rs.values, name=f"{idx}", mode="lines+markers",
                    line=dict(width=2),
                    marker=dict(symbol="asterisk", size=8, angleref="previous")
                ))
            st.plotly_chart(fig, use_container_width=True)
            # st.altair_chart(Ui.create_altair_m_line_plot_with(df_unique_values), use_container_width=True)

            with st.expander(f"Post-opt Groups", expanded=True):
                for i in df_col_group_dict:
                    st.dataframe(pd.DataFrame([list(df_col_group_dict[i].columns)], index=[i]))
                st.write(df_unique_values)


def multiline_chart_from_df_columns(df: pd.DataFrame, apply_func=None, x_name=" ", y_name="value", **kwargs):
    if apply_func is not None:
        df = df.apply(apply_func)
    df = df.reset_index().melt(id_vars=["index"])
    return px.line(
        df, color="variable", y="value", x="index", labels={'index': x_name, 'value': y_name}, **kwargs
    )


# ----------------
# "https://github.com/whitphx/streamlit-theme-editor/blob/main/app.py"

@st.cache_resource
def get_config_theme_color():
    config_theme_primary_color = st._config.get_option('theme.primaryColor')
    config_theme_background_color = st._config.get_option('theme.backgroundColor')
    config_theme_secondary_background_color = st._config.get_option('theme.secondaryBackgroundColor')
    config_theme_text_color = st._config.get_option('theme.textColor')
    if config_theme_primary_color and config_theme_background_color and config_theme_secondary_background_color and config_theme_text_color:
        return dict(
            primaryColor=config_theme_primary_color,
            backgroundColor=config_theme_background_color,
            secondaryBackgroundColor=config_theme_secondary_background_color,
            textColor=config_theme_text_color,
        )

    return None


def color_theme_main():
    if "preset_colors" not in st.session_state:
        st.session_state["preset_colors"] = {
            "Default dark": dict(
                primaryColor="#ff4b4b",
                backgroundColor="#0e1117",
                secondaryBackgroundColor="#262730",
                textColor="#fafafa",
            ),
            "Default light": dict(
                primaryColor="#ff4b4b",
                backgroundColor="#ffffff",
                secondaryBackgroundColor="#f0f2f6",
                textColor="#31333F",
            ),
            "Pandas": dict(
                primaryColor="#3FB1C5",
                backgroundColor="#14181E",
                secondaryBackgroundColor="#29313D",
                textColor="#CED6DD",
            ),
            "Green Gold": dict(
                primaryColor="#A6D645",
                backgroundColor="#444A41",
                secondaryBackgroundColor="#252521",
                textColor="#EBEBFF",
            ),
            "scipy": dict(
                primaryColor="#459DB9",
                backgroundColor="#121212",
                secondaryBackgroundColor="#1E1E1E",
                textColor="#C9D1D9",
            ),
            "youtube": dict(
                primaryColor="#FF0000",
                backgroundColor="#0F0F0F",
                secondaryBackgroundColor="#272727",
                textColor="#F1F1F1",
            ),
            "github": dict(
                primaryColor="#F78166",
                backgroundColor="#0D1117",
                secondaryBackgroundColor="#161B22",
                textColor="#E6EDF3",
            ),
            "old_car_panel_green": dict(
                primaryColor="#15A819",  # 00E282 0BD87E 00A841, 14C128, 15A819, 00A841, 15A834
                backgroundColor="#020402",
                secondaryBackgroundColor="#17201E",     # 212525 1A1E1E 17201D
                textColor="#DEEAD5",
            ),
            "darker_theme": dict(
                primaryColor="#E82736",     # DA213B
                backgroundColor="#080808",
                secondaryBackgroundColor="#1A1A18",
                textColor="#F1EDED",
            ),
        }

    preset_colors = st.session_state["preset_colors"]
    default_color = None
    theme_from_initial_config = get_config_theme_color()
    if theme_from_initial_config:
        preset_colors["From_the_config"] = theme_from_initial_config
        default_color = preset_colors["From_the_config"]

    def st_slt_box(options):
        return st.selectbox("Select theme", options=options, index=0)
    default_color = preset_colors[st_slt_box(preset_colors.keys())]

    theme_name = st.text_input("", placeholder="Name to save theme", label_visibility="hidden")
    if theme_name != "" and theme_name is not None and theme_name != " " and theme_name not in preset_colors:
        print("line382")
        print(preset_colors.keys())
        print(theme_name)
        st.session_state["preset_colors"][theme_name] = dict(
            primaryColor=st.session_state["primaryColor"],
            backgroundColor=st.session_state["backgroundColor"],
            secondaryBackgroundColor=st.session_state["secondaryBackgroundColor"],
            textColor=st.session_state["textColor"],
        )
        print(st.session_state["preset_colors"].keys())

    for key in ('primaryColor', 'backgroundColor', 'secondaryBackgroundColor', 'textColor'):
        color = st.color_picker(f"{key}", key=f"st_widget{key}", value=default_color[key])
        st.session_state[key] = color

    if st.checkbox("Apply theme to this page"):
        st.info("Select 'Custom Theme' in the settings dialog to see the effect")
        for key in ('primaryColor', 'backgroundColor', 'secondaryBackgroundColor', 'textColor'):
            print(st.session_state[key])
        apply_theme()


def apply_theme():
    def reconcile_theme_config():
        keys = ['primaryColor', 'backgroundColor', 'secondaryBackgroundColor', 'textColor']
        has_changed = False
        for key in keys:
            if st._config.get_option(f'theme.{key}') != st.session_state[key]:
                st._config.set_option(f'theme.{key}', st.session_state[key])
                has_changed = True
        if has_changed:
            st.rerun()

    reconcile_theme_config()
